import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os

class EmailSenderGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Email Sender with PDF Attachments")
        self.root.geometry("700x600")
        
        self.excel_file = None
        self.recipients = []
        self.workbook = None
        self.sheet_names = []
        
        # Excel File Selection
        tk.Label(root, text="Excel File:", font=("Arial", 10, "bold")).pack(pady=(10, 5))
        file_frame = tk.Frame(root)
        file_frame.pack(pady=5)
        
        self.file_label = tk.Label(file_frame, text="No file selected", fg="gray")
        self.file_label.pack(side=tk.LEFT, padx=5)
        
        tk.Button(file_frame, text="Browse Excel", command=self.load_excel).pack(side=tk.LEFT)
        
        # Sheet Selection
        sheet_frame = tk.Frame(root)
        sheet_frame.pack(pady=5)
        
        tk.Label(sheet_frame, text="Select Sheet:").pack(side=tk.LEFT, padx=5)
        self.sheet_var = tk.StringVar()
        self.sheet_dropdown = tk.OptionMenu(sheet_frame, self.sheet_var, "")
        self.sheet_dropdown.config(state=tk.DISABLED, width=20)
        self.sheet_dropdown.pack(side=tk.LEFT, padx=5)
        
        tk.Button(sheet_frame, text="Load Sheet", command=self.load_sheet, state=tk.DISABLED).pack(side=tk.LEFT)
        self.load_sheet_button = sheet_frame.winfo_children()[-1]
        
        # Gmail Credentials
        tk.Label(root, text="Gmail Account:", font=("Arial", 10, "bold")).pack(pady=(10, 5))
        
        cred_frame = tk.Frame(root)
        cred_frame.pack(pady=5)
        
        tk.Label(cred_frame, text="Email:").grid(row=0, column=0, padx=5, sticky="e")
        self.email_entry = tk.Entry(cred_frame, width=30)
        self.email_entry.grid(row=0, column=1, padx=5)
        
        tk.Label(cred_frame, text="App Password:").grid(row=1, column=0, padx=5, sticky="e")
        self.password_entry = tk.Entry(cred_frame, width=30, show="*")
        self.password_entry.grid(row=1, column=1, padx=5)
        
        # Subject
        tk.Label(root, text="Email Subject:", font=("Arial", 10, "bold")).pack(pady=(10, 5))
        self.subject_entry = tk.Entry(root, width=60)
        self.subject_entry.pack(pady=5)
        
        # Message Body
        tk.Label(root, text="Email Message (use {name} to insert recipient's name):", font=("Arial", 10, "bold")).pack(pady=(10, 5))
        self.message_text = scrolledtext.ScrolledText(root, width=70, height=10)
        self.message_text.pack(pady=5)
        self.message_text.insert("1.0", "Dear {name},\n\nYour message here...")
        
        # Send Button
        self.send_button = tk.Button(root, text="Send Emails", command=self.send_emails, 
                                     bg="#4CAF50", fg="white", font=("Arial", 12, "bold"))
        self.send_button.pack(pady=10)
        
        # Status Log
        tk.Label(root, text="Status Log:", font=("Arial", 10, "bold")).pack(pady=(10, 5))
        self.log_text = scrolledtext.ScrolledText(root, width=70, height=8)
        self.log_text.pack(pady=5)
        
    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update()
        
    def load_excel(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if not file_path:
            return
            
        try:
            self.workbook = openpyxl.load_workbook(file_path)
            self.sheet_names = self.workbook.sheetnames
            self.excel_file = file_path
            
            # Update dropdown with sheet names
            menu = self.sheet_dropdown["menu"]
            menu.delete(0, "end")
            for sheet_name in self.sheet_names:
                menu.add_command(label=sheet_name, 
                               command=lambda value=sheet_name: self.sheet_var.set(value))
            
            # Set default to first sheet
            self.sheet_var.set(self.sheet_names[0])
            self.sheet_dropdown.config(state=tk.NORMAL)
            self.load_sheet_button.config(state=tk.NORMAL)
            
            self.file_label.config(text=f"Loaded: {os.path.basename(file_path)} ({len(self.sheet_names)} sheets)", fg="blue")
            self.log(f"Excel file loaded: {os.path.basename(file_path)}")
            self.log(f"Available sheets: {', '.join(self.sheet_names)}")
            self.log("Please select a sheet and click 'Load Sheet'")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel file:\n{str(e)}")
    
    def load_sheet(self):
        selected_sheet = self.sheet_var.get()
        
        if not selected_sheet:
            messagebox.showwarning("Warning", "Please select a sheet!")
            return
            
        try:
            sheet = self.workbook[selected_sheet]
            self.recipients = []
            
            # Assume first row is header
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] and row[2]:  # Name, Email, AttachmentPath
                    self.recipients.append({
                        'name': row[0],
                        'email': row[1],
                        'attachment': row[2]
                    })
            
            self.file_label.config(text=f"Sheet '{selected_sheet}': {len(self.recipients)} recipients loaded", fg="green")
            self.log(f"Loaded {len(self.recipients)} recipients from sheet '{selected_sheet}'")
            
            if len(self.recipients) == 0:
                messagebox.showwarning("Warning", "No recipients found in this sheet!\n\nMake sure the sheet has:\n- Row 1: Headers (Name, Email, AttachmentPath)\n- Row 2+: Data")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sheet:\n{str(e)}")
            
    def send_emails(self):
        # Validate inputs
        if not self.recipients:
            messagebox.showwarning("Warning", "Please load an Excel file first!")
            return
            
        sender_email = self.email_entry.get().strip()
        sender_password = self.password_entry.get().strip()
        subject = self.subject_entry.get().strip()
        message_body = self.message_text.get("1.0", tk.END).strip()
        
        if not sender_email or not sender_password:
            messagebox.showwarning("Warning", "Please enter Gmail credentials!")
            return
            
        if not subject or not message_body:
            messagebox.showwarning("Warning", "Please enter subject and message!")
            return
        
        # Confirm before sending
        if not messagebox.askyesno("Confirm", f"Send {len(self.recipients)} emails?"):
            return
            
        self.send_button.config(state=tk.DISABLED)
        self.log("Starting email sending process...")
        
        success_count = 0
        fail_count = 0
        
        try:
            # Connect to Gmail SMTP server
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(sender_email, sender_password)
            self.log("Connected to Gmail successfully")
            
            for recipient in self.recipients:
                try:
                    # Create message
                    msg = MIMEMultipart()
                    msg['From'] = sender_email
                    msg['To'] = recipient['email']
                    msg['Subject'] = subject
                    
                    # Personalize message with name (replace {name} placeholder)
                    personalized_message = message_body.replace("{name}", recipient['name'])
                    msg.attach(MIMEText(personalized_message, 'plain'))
                    
                    # Attach PDF
                    if os.path.exists(recipient['attachment']):
                        with open(recipient['attachment'], 'rb') as attachment:
                            part = MIMEBase('application', 'octet-stream')
                            part.set_payload(attachment.read())
                            encoders.encode_base64(part)
                            part.add_header(
                                'Content-Disposition',
                                f'attachment; filename= {os.path.basename(recipient["attachment"])}'
                            )
                            msg.attach(part)
                    else:
                        self.log(f"⚠ Warning: Attachment not found for {recipient['name']}: {recipient['attachment']}")
                    
                    # Send email
                    server.send_message(msg)
                    success_count += 1
                    self.log(f"✓ Sent to {recipient['name']} ({recipient['email']})")
                    
                except Exception as e:
                    fail_count += 1
                    self.log(f"✗ Failed to send to {recipient['name']}: {str(e)}")
            
            server.quit()
            self.log(f"\nCompleted! Success: {success_count}, Failed: {fail_count}")
            messagebox.showinfo("Complete", f"Emails sent!\nSuccess: {success_count}\nFailed: {fail_count}")
            
        except Exception as e:
            self.log(f"Error: {str(e)}")
            messagebox.showerror("Error", f"Failed to connect to Gmail:\n{str(e)}\n\nMake sure you're using an App Password, not your regular password.")
        
        finally:
            self.send_button.config(state=tk.NORMAL)

if __name__ == "__main__":
    root = tk.Tk()
    app = EmailSenderGUI(root)
    root.mainloop()
