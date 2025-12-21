import os
from datetime import datetime, timedelta

# Force GTK to use your system theme for file dialogs - must be set before importing tkinter
os.environ['GTK_THEME'] = 'Qogir-Dark'
os.environ['GTK2_RC_FILES'] = '/usr/share/themes/Qogir-Dark/gtk-2.0/gtkrc'

import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox, ttk
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

class RoundedEntry(tk.Canvas):
    def __init__(self, parent, width=200, height=30, corner_radius=15, bg_color='#9da9a0', 
                 fg_color='#232a2e', border_color='#425047', font=("MesloLGS Nerd Font", 10), **kwargs):
        tk.Canvas.__init__(self, parent, width=width, height=height, bg=parent['bg'], 
                          highlightthickness=0, **kwargs)
        self.corner_radius = corner_radius
        self.bg_color = bg_color
        self.border_color = border_color
        
        # Draw rounded rectangle
        self.create_rounded_rect(2, 2, width-2, height-2, corner_radius, 
                                fill=bg_color, outline=border_color, width=2)
        
        # Create entry widget
        self.entry = tk.Entry(self, bg=bg_color, fg=fg_color, relief='flat', 
                             bd=0, highlightthickness=0, font=font)
        self.create_window(width/2, height/2, window=self.entry, width=width-20, height=height-10)
    
    def create_rounded_rect(self, x1, y1, x2, y2, radius, **kwargs):
        points = [x1+radius, y1,
                 x2-radius, y1,
                 x2, y1,
                 x2, y1+radius,
                 x2, y2-radius,
                 x2, y2,
                 x2-radius, y2,
                 x1+radius, y2,
                 x1, y2,
                 x1, y2-radius,
                 x1, y1+radius,
                 x1, y1]
        return self.create_polygon(points, smooth=True, **kwargs)
    
    def get(self):
        return self.entry.get()
    
    def insert(self, index, string):
        return self.entry.insert(index, string)
    
    def delete(self, first, last=None):
        return self.entry.delete(first, last)
    
    def config_entry(self, **kwargs):
        self.entry.config(**kwargs)

class RoundedText(tk.Frame):
    def __init__(self, parent, width=500, height=150, corner_radius=15, bg_color='#9da9a0', 
                 fg_color='#232a2e', border_color='#425047', **kwargs):
        tk.Frame.__init__(self, parent, bg=parent['bg'])
        
        # Canvas for rounded border
        self.canvas = tk.Canvas(self, width=width, height=height, bg=parent['bg'], 
                               highlightthickness=0)
        self.canvas.pack()
        
        # Draw rounded rectangle
        self.create_rounded_rect(self.canvas, 2, 2, width-2, height-2, corner_radius, 
                                fill=bg_color, outline=border_color, width=2)
        
        # Create text widget
        self.text = tk.Text(self.canvas, bg=bg_color, fg=fg_color, relief='flat', 
                           bd=0, highlightthickness=0, **kwargs)
        self.canvas.create_window(width/2, height/2, window=self.text, 
                                 width=width-20, height=height-10)
    
    def create_rounded_rect(self, canvas, x1, y1, x2, y2, radius, **kwargs):
        points = [x1+radius, y1,
                 x2-radius, y1,
                 x2, y1,
                 x2, y1+radius,
                 x2, y2-radius,
                 x2, y2,
                 x2-radius, y2,
                 x1+radius, y2,
                 x1, y2,
                 x1, y2-radius,
                 x1, y1+radius,
                 x1, y1]
        return canvas.create_polygon(points, smooth=True, **kwargs)
    
    def get(self, index1, index2=None):
        return self.text.get(index1, index2)
    
    def insert(self, index, string):
        return self.text.insert(index, string)
    
    def delete(self, index1, index2=None):
        return self.text.delete(index1, index2)
    
    def config_text(self, **kwargs):
        self.text.config(**kwargs)

class RoundedButton(tk.Canvas):
    def __init__(self, parent, text="Button", command=None, width=120, height=40, 
                 corner_radius=15, bg_color='#a7c080', fg_color='#2d353b', 
                 hover_color='#83c092', border_color='#425047', font=("MesloLGS Nerd Font", 10), **kwargs):
        tk.Canvas.__init__(self, parent, width=width, height=height, bg=parent['bg'], 
                          highlightthickness=0, **kwargs)
        self.corner_radius = corner_radius
        self.bg_color = bg_color
        self.hover_color = hover_color
        self.border_color = border_color
        self.command = command
        self.is_enabled = True
        
        # Draw rounded rectangle
        self.rect = self.create_rounded_rect(2, 2, width-2, height-2, corner_radius, 
                                             fill=bg_color, outline=border_color, width=2)
        
        # Create text
        self.text_id = self.create_text(width/2, height/2, text=text, fill=fg_color, font=font)
        
        # Bind events
        self.bind("<Button-1>", self._on_click)
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)
    
    def create_rounded_rect(self, x1, y1, x2, y2, radius, **kwargs):
        points = [x1+radius, y1,
                 x2-radius, y1,
                 x2, y1,
                 x2, y1+radius,
                 x2, y2-radius,
                 x2, y2,
                 x2-radius, y2,
                 x1+radius, y2,
                 x1, y2,
                 x1, y2-radius,
                 x1, y1+radius,
                 x1, y1]
        return self.create_polygon(points, smooth=True, **kwargs)
    
    def _on_click(self, event):
        if self.command and self.is_enabled:
            self.command()
    
    def _on_enter(self, event):
        if self.is_enabled:
            self.itemconfig(self.rect, fill=self.hover_color)
    
    def _on_leave(self, event):
        if self.is_enabled:
            self.itemconfig(self.rect, fill=self.bg_color)
    
    def config_button(self, state=None, bg=None, **kwargs):
        if state == tk.DISABLED or state == 'disabled':
            self.is_enabled = False
            self.itemconfig(self.rect, fill='#232a2e')
            self.itemconfig(self.text_id, fill='#56635f')
        elif state == tk.NORMAL or state == 'normal':
            self.is_enabled = True
            self.itemconfig(self.rect, fill=self.bg_color)
            self.itemconfig(self.text_id, fill='#2d353b')
        if bg:
            self.bg_color = bg
            if self.is_enabled:
                self.itemconfig(self.rect, fill=bg)

class RoundedDropdown(tk.Canvas):
    def __init__(self, parent, variable, values=None, width=200, height=35, 
                 corner_radius=15, bg_color='#7fbbb3', fg_color='#a7c080',
                 hover_color='#83c092', border_color='#425047', 
                 arrow_color_disabled='#9da9d0', arrow_color_enabled='#83c092', **kwargs):
        tk.Canvas.__init__(self, parent, width=width, height=height, bg=parent['bg'], 
                          highlightthickness=0, **kwargs)
        self.corner_radius = corner_radius
        self.bg_color = bg_color
        self.hover_color = hover_color
        self.border_color = border_color
        self.arrow_color_disabled = arrow_color_disabled
        self.arrow_color_enabled = arrow_color_enabled
        self.variable = variable
        self.values = values if values else []
        self.is_enabled = True
        self.menu_window = None
        
        # Draw rounded rectangle
        self.rect = self.create_rounded_rect(2, 2, width-2, height-2, corner_radius, 
                                             fill=bg_color, outline=border_color, width=2)
        
        # Create text
        self.text_id = self.create_text(20, height/2, text="", fill=fg_color, 
                                       font=("MesloLGS Nerd Font", 10, "bold"), anchor='w')
        
        # Create arrow
        arrow_x = width - 20
        arrow_y = height / 2
        self.arrow = self.create_polygon(
            arrow_x - 5, arrow_y - 3,
            arrow_x + 5, arrow_y - 3,
            arrow_x, arrow_y + 4,
            fill=fg_color, outline=fg_color
        )
        
        # Bind events
        self.bind("<Button-1>", self._on_click)
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)
        
        # Update text when variable changes
        if self.variable:
            self.variable.trace('w', self._update_text)
            self._update_text()
    
    def create_rounded_rect(self, x1, y1, x2, y2, radius, **kwargs):
        points = [x1+radius, y1,
                 x2-radius, y1,
                 x2, y1,
                 x2, y1+radius,
                 x2, y2-radius,
                 x2, y2,
                 x2-radius, y2,
                 x1+radius, y2,
                 x1, y2,
                 x1, y2-radius,
                 x1, y1+radius,
                 x1, y1]
        return self.create_polygon(points, smooth=True, **kwargs)
    
    def _update_text(self, *args):
        if self.variable:
            text = self.variable.get()
            if len(text) > 25:
                text = text[:22] + "..."
            self.itemconfig(self.text_id, text=text)
    
    def _on_click(self, event):
        if not self.is_enabled or not self.values:
            return
        
        # Create popup menu
        menu = tk.Menu(self, tearoff=0, bg=self.bg_color, fg='#1e293b',
                      activebackground=self.hover_color, activeforeground='#1e293b',
                      borderwidth=2, relief='solid', font=("MesloLGS Nerd Font", 10, "bold"))
        
        for value in self.values:
            menu.add_command(label=value, 
                           command=lambda v=value: self.variable.set(v))
        
        # Show menu below the dropdown
        x = self.winfo_rootx()
        y = self.winfo_rooty() + self.winfo_height()
        menu.post(x, y)
    
    def _on_enter(self, event):
        if self.is_enabled:
            self.itemconfig(self.rect, fill=self.hover_color)
    
    def _on_leave(self, event):
        if self.is_enabled:
            self.itemconfig(self.rect, fill=self.bg_color)
    
    def set_values(self, values):
        self.values = values
    
    def config_dropdown(self, state=None, **kwargs):
        if state == tk.DISABLED or state == 'disabled':
            self.is_enabled = False
            self.itemconfig(self.rect, fill='#3a515d')
            self.itemconfig(self.text_id, fill='#56635f')
            self.itemconfig(self.arrow, fill=self.arrow_color_disabled)
        elif state == tk.NORMAL or state == 'normal':
            self.is_enabled = True
            self.itemconfig(self.rect, fill=self.bg_color)
            self.itemconfig(self.text_id, fill='#1e293b')
            self.itemconfig(self.arrow, fill=self.arrow_color_enabled)

class EmailSenderGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Email Sender with PDF Attachments")
        self.root.geometry("700x700")
        
        # Everforest Medium Dark Color Scheme
        self.colors = {
            'window_bg': '#2d353b',           # bg0 - Default Background
            'window_fg': '#d3c6aa',           # fg - Default Foreground
            'send_btn_bg': '#859966',         # green - Primary action
            'send_btn_fg': '#1E293B',         # bg0 - Text on buttons
            'browse_excel_btn_bg': '#859966', # green - Primary action
            'browse_excel_btn_fg': '#1E293B', # bg0 - Text on buttons
            'dropdown_bg': '#65958f',         # blue - Interactive element
            'dropdown_fg': '#2d353b',         # bg0 - Text on dropdown
            'load_sheet_btn_bg_active': '#859966',    # green - Active state
            'load_sheet_btn_fg_active': '#1E293B',    # bg0 - Text
            'load_sheet_btn_bg_inactive': '#232a2e',  # bg_dim - Inactive state
            'load_sheet_btn_fg_inactive': '#859289',  # grey1 - Disabled text
            'field_bg': '#7a8478',            # bg2 - Input fields
            'field_fg': '#1E293B',            # fg - Input text
            'status_log_bg': '#7a8478',       # bg1 - Status area
            'status_log_fg': '#1E293B',       # fg - Log text
            'titles_fg': '#d699b6',           # purple - Section titles
            'labels_fg': '#dbbc7f',           # yellow - Form labels
            'hover_bg': '#689974',            # aqua - Hover state
            'not_loaded_excel_fg': '#859289', # grey1 - Muted text
            'loaded_excel_fg': '#859966',     # green - Success state
            'scrollbar_bg': '#475258',        # bg3 - Scrollbar
            'field_border': '#425047',         # bg_green - Borders
            'hover_bg_2':   '#9da9d0',           # light-grey - Hover state
        }
        
        # Set window background color
        self.root.config(bg=self.colors['window_bg'])
        
        # Configure scrollbar style - force custom colors
        style = ttk.Style()
        style.theme_use('clam')  # Use 'clam' theme which allows better customization
        style.configure("Vertical.TScrollbar", 
                       background=self.colors['scrollbar_bg'],
                       troughcolor=self.colors['window_bg'],
                       bordercolor=self.colors['window_bg'],
                       arrowcolor=self.colors['window_fg'],
                       darkcolor=self.colors['scrollbar_bg'],
                       lightcolor=self.colors['scrollbar_bg'])
        style.map("Vertical.TScrollbar",
                 background=[('active', self.colors['hover_bg']), 
                           ('!active', self.colors['scrollbar_bg'])])
        
        self.excel_file = None
        self.recipients = []
        self.workbook = None
        self.sheet_names = []
        
        # Excel File Selection
        tk.Label(root, text="Excel File:", font=("MesloLGS Nerd Font", 10, "bold"), 
                fg=self.colors['titles_fg'], bg=self.colors['window_bg']).pack(pady=(10, 5))
        file_frame = tk.Frame(root)
        file_frame.pack(pady=5)
        file_frame.config(bg=self.colors['window_bg'])
        
        self.file_label = tk.Label(file_frame, text="No file selected", 
                                   fg=self.colors['not_loaded_excel_fg'], bg=self.colors['window_bg'])
        self.file_label.pack(side=tk.LEFT, padx=5)
        
        self.browse_btn = RoundedButton(file_frame, text="Browse Excel", command=self.load_excel,
                                        width=140, height=35, corner_radius=15,
                                        bg_color=self.colors['browse_excel_btn_bg'],
                                        fg_color=self.colors['browse_excel_btn_fg'],
                                        hover_color=self.colors['hover_bg'],
                                        border_color=self.colors['field_border'],
                                        font=("MesloLGS Nerd Font", 10, "bold"))
        self.browse_btn.pack(side=tk.LEFT)
        
        # Sheet Selection
        sheet_frame = tk.Frame(root)
        sheet_frame.pack(pady=5)
        sheet_frame.config(bg=self.colors['window_bg'])
        
        tk.Label(sheet_frame, text="Select Sheet:", 
                fg=self.colors['titles_fg'], bg=self.colors['window_bg']).pack(side=tk.LEFT, padx=5)
        self.sheet_var = tk.StringVar()
        self.sheet_dropdown = RoundedDropdown(sheet_frame, self.sheet_var, values=[],
                                              width=180, height=35, corner_radius=15,
                                              bg_color=self.colors['dropdown_bg'],
                                              fg_color=self.colors['dropdown_fg'],
                                              hover_color=self.colors['hover_bg'],
                                              border_color=self.colors['field_border'],
                                              arrow_color_disabled=self.colors['hover_bg_2'],
                                              arrow_color_enabled=self.colors['send_btn_fg'])
        self.sheet_dropdown.config_dropdown(state=tk.DISABLED)
        self.sheet_dropdown.pack(side=tk.LEFT, padx=5)
        
        self.load_sheet_button = RoundedButton(sheet_frame, text="Load Sheet", command=self.load_sheet,
                                               width=120, height=35, corner_radius=15,
                                               bg_color=self.colors['load_sheet_btn_bg_inactive'],
                                               fg_color=self.colors['load_sheet_btn_fg_inactive'],
                                               hover_color=self.colors['hover_bg'],
                                               border_color=self.colors['field_border'],
                                               font=("MesloLGS Nerd Font", 10, "bold"))
        self.load_sheet_button.config_button(state=tk.DISABLED)
        self.load_sheet_button.pack(side=tk.LEFT)
        
        # Gmail Credentials
        tk.Label(root, text="Gmail Account:", font=("MesloLGS Nerd Font", 10, "bold"), 
                fg=self.colors['titles_fg'], bg=self.colors['window_bg']).pack(pady=(10, 5))
        
        cred_frame = tk.Frame(root)
        cred_frame.pack(pady=5)
        cred_frame.config(bg=self.colors['window_bg'])
        
        tk.Label(cred_frame, text="Email:", 
                fg='#d699b6', bg=self.colors['window_bg']).grid(row=0, column=0, padx=5, sticky="e")
        self.email_entry = RoundedEntry(cred_frame, width=240, height=30, corner_radius=15,
                                       bg_color=self.colors['field_bg'], 
                                       fg_color=self.colors['field_fg'],
                                       border_color=self.colors['field_border'],
                                       font=("MesloLGS Nerd Font", 10, "bold"))
        self.email_entry.grid(row=0, column=1, padx=5, pady=5)
        
        tk.Label(cred_frame, text="App Password:", 
                fg='#d699b6', bg=self.colors['window_bg']).grid(row=1, column=0, padx=5, sticky="e")
        self.password_entry = RoundedEntry(cred_frame, width=240, height=30, corner_radius=15,
                                          bg_color=self.colors['field_bg'], 
                                          fg_color=self.colors['field_fg'],
                                          border_color=self.colors['field_border'],
                                          font=("MesloLGS Nerd Font", 10, "bold"))
        self.password_entry.config_entry(show="*")
        self.password_entry.grid(row=1, column=1, padx=5, pady=5)
        
        # Subject
        tk.Label(root, text="Email Subject:", font=("MesloLGS Nerd Font", 10, "bold"), 
                fg=self.colors['titles_fg'], bg=self.colors['window_bg']).pack(pady=(10, 5))
        self.subject_entry = RoundedEntry(root, width=480, height=30, corner_radius=15,
                                         bg_color=self.colors['field_bg'], 
                                         fg_color=self.colors['field_fg'],
                                         border_color=self.colors['field_border'])
        self.subject_entry.pack(pady=5)
        # Calculate previous month
        today = datetime.today()
        first_day_this_month = today.replace(day=1)
        last_month = first_day_this_month - timedelta(days=1)
        month_year = last_month.strftime("%m.%Y")
        self.subject_entry.insert(0, f"Ливче од пресметка на плата за {month_year}")
        
        # Message Body
        tk.Label(root, text="Email Message (use {name} to insert recipient's name):", 
                font=("MesloLGS Nerd Font", 10, "bold"), fg=self.colors['titles_fg'], 
                bg=self.colors['window_bg']).pack(pady=(10, 5))
        
        # Create rounded text widget for message
        self.message_text = RoundedText(root, width=560, height=150, corner_radius=15,
                                       bg_color=self.colors['field_bg'], 
                                       fg_color=self.colors['field_fg'],
                                       border_color=self.colors['field_border'],
                                       wrap=tk.WORD)
        self.message_text.pack(pady=5)
        
        self.message_text.insert("1.0", f"Почитуван {{name}},\n\n\nВо прилог ти праќам ливче од пресметка на плата за {month_year}\n\n\nР.Ѕ. Доколку воочите дека има нејаснотии во ливчето задолжително јавете се во Благајна или на емаил takidaskalo.blagajnik@gmail.com")
        
        # Send Button
        self.send_button = RoundedButton(root, text="Send Emails", command=self.send_emails,
                                        width=180, height=50, corner_radius=20,
                                        bg_color=self.colors['send_btn_bg'],
                                        fg_color=self.colors['send_btn_fg'],
                                        hover_color=self.colors['hover_bg'],
                                        border_color=self.colors['field_border'],
                                        font=("MesloLGS Nerd Font", 12, "bold"))
        self.send_button.pack(pady=10)
        
        # Status Log
        tk.Label(root, text="Status Log:", font=("MesloLGS Nerd Font", 10, "bold"), 
                fg=self.colors['titles_fg'], bg=self.colors['window_bg']).pack(pady=(10, 5))
        
        # Create rounded text widget for status log
        self.log_text = RoundedText(root, width=560, height=180, corner_radius=15,
                                   bg_color=self.colors['status_log_bg'], 
                                   fg_color=self.colors['status_log_fg'],
                                   border_color=self.colors['field_border'],
                                   wrap=tk.WORD)
        self.log_text.pack(pady=5)
        
    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.text.see(tk.END)
        self.root.update()
    
    def load_excel(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if not file_path:
            return
            
        try:
            # FIX: Added data_only=True to read calculated values instead of formulas
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            self.sheet_names = self.workbook.sheetnames
            self.excel_file = file_path
            
            # Update dropdown with sheet names
            self.sheet_dropdown.set_values(self.sheet_names)
            
            # Set default to first sheet
            self.sheet_var.set(self.sheet_names[0])
            self.sheet_dropdown.config_dropdown(state=tk.NORMAL)
            self.load_sheet_button.bg_color = self.colors['load_sheet_btn_bg_active']
            self.load_sheet_button.config_button(state=tk.NORMAL)
            
            self.file_label.config(text=f"Loaded: {os.path.basename(file_path)} ({len(self.sheet_names)} sheets)", 
                                  fg=self.colors['loaded_excel_fg'])
            self.log(f"Excel file loaded: {os.path.basename(file_path)}")
            self.log(f"Available sheets: {', '.join(self.sheet_names)}")
            self.log("Please select a sheet and click 'Load Sheet'")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel file:\n{str(e)}")
    
    def load_sheet(self):
        selected_sheet = self.sheet_var.get()
        
        if not selected_sheet:
            messagebox.showwarning("Warning", "Please select a sheet!")
            return
            
        try:
            sheet = self.workbook[selected_sheet]
            self.recipients = []
            
            # Assume first row is header
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] and row[2]:  # Name, Email, AttachmentPath
                    self.recipients.append({
                        'name': row[0],
                        'email': row[1],
                        'attachment': row[2]
                    })
            
            self.file_label.config(text=f"Sheet '{selected_sheet}': {len(self.recipients)} recipients loaded", 
                                  fg=self.colors['loaded_excel_fg'])
            self.log(f"Loaded {len(self.recipients)} recipients from sheet '{selected_sheet}'")
            
            if len(self.recipients) == 0:
                messagebox.showwarning("Warning", "No recipients found in this sheet!\n\nMake sure the sheet has:\n- Row 1: Headers (Name, Email, AttachmentPath)\n- Row 2+: Data")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sheet:\n{str(e)}")
            
    def send_emails(self):
        # Validate inputs
        if not self.recipients:
            messagebox.showwarning("Warning", "Please load an Excel file first!")
            return
            
        sender_email = self.email_entry.get().strip()
        sender_password = self.password_entry.get().strip()
        subject = self.subject_entry.get().strip()
        message_body = self.message_text.get("1.0", tk.END).strip()
        
        if not sender_email or not sender_password:
            messagebox.showwarning("Warning", "Please enter Gmail credentials!")
            return
            
        if not subject or not message_body:
            messagebox.showwarning("Warning", "Please enter subject and message!")
            return
        
        # Confirm before sending
        if not messagebox.askyesno("Confirm", f"Send {len(self.recipients)} emails?"):
            return
            
        self.send_button.config_button(state=tk.DISABLED)
        self.log("Starting email sending process...")
        
        success_count = 0
        fail_count = 0
        
        try:
            # Connect to Gmail SMTP server
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(sender_email, sender_password)
            self.log("Connected to Gmail successfully")
            
            for recipient in self.recipients:
                try:
                    # Create message
                    msg = MIMEMultipart()
                    msg['From'] = sender_email
                    msg['To'] = recipient['email']
                    msg['Subject'] = subject
                    
                    # Personalize message with name (replace {name} placeholder)
                    personalized_message = message_body.replace("{name}", recipient['name'])
                    msg.attach(MIMEText(personalized_message, 'plain'))
                    
                    # Attach PDF
                    if os.path.exists(recipient['attachment']):
                        attachment_path = recipient['attachment']
                        filename = os.path.basename(attachment_path)
                        with open(attachment_path, 'rb') as file:
                            part = MIMEBase('application', 'octet-stream')
                            part.set_payload(file.read())
                            encoders.encode_base64(part)
                            part.add_header('Content-Disposition', 'attachment', filename=filename)
                            msg.attach(part)
                    else:
                        self.log(f"⚠ Warning: Attachment not found for {recipient['name']}: {recipient['attachment']}")
                    
                    # Send email
                    server.send_message(msg)
                    success_count += 1
                    self.log(f"✓ Sent to {recipient['name']} ({recipient['email']})")
                    
                except Exception as e:
                    fail_count += 1
                    self.log(f"✗ Failed to send to {recipient['name']}: {str(e)}")
            
            server.quit()
            self.log(f"\nCompleted! Success: {success_count}, Failed: {fail_count}")
            messagebox.showinfo("Complete", f"Emails sent!\nSuccess: {success_count}\nFailed: {fail_count}")
            
        except Exception as e:
            self.log(f"Error: {str(e)}")
            messagebox.showerror("Error", f"Failed to connect to Gmail:\n{str(e)}\n\nMake sure you're using an App Password, not your regular password.")
        
        finally:
            self.send_button.config_button(state=tk.NORMAL)

if __name__ == "__main__":
    root = tk.Tk()
    app = EmailSenderGUI(root)
    root.mainloop()
