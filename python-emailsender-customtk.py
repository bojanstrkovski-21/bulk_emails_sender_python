import customtkinter as ctk
from tkinter import filedialog, messagebox
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
from datetime import datetime, timedelta

class EmailSenderApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Email Sender with PDF Attachments")
        self.geometry("700x700")
        ctk.set_appearance_mode("dark")
        # Removed set_default_color_theme to allow full custom color control

        # Everforest Medium Dark Color Scheme
        self.colors = {
            'window_bg': '#2d353b',           # bg0 - Default Background
            'window_fg': '#d3c6aa',           # fg - Default Foreground
            'send_btn_bg': '#859966',         # green - Primary action
            'send_btn_fg': '#1E293B',         # bg0 - Text on buttons
            'browse_excel_btn_bg': '#859966', # green - Primary action
            'browse_excel_btn_fg': '#1E293B', # bg0 - Text on buttons
            'dropdown_bg': '#65958f',         # blue - Interactive element
            'dropdown_fg': '#2d353b',         # bg0 - Text on dropdown
            'load_sheet_btn_bg_active': '#859966',    # green - Active state
            'load_sheet_btn_fg_active': '#1E293B',    # bg0 - Text
            'load_sheet_btn_bg_inactive': '#232a2e',  # bg_dim - Inactive state
            'load_sheet_btn_fg_inactive': '#859289',  # grey1 - Disabled text
            'field_bg': '#7a8478',            # bg2 - Input fields
            'field_fg': '#1E293B',            # fg - Input text
            'status_log_bg': '#7a8478',       # bg1 - Status area
            'status_log_fg': '#1E293B',       # fg - Log text
            'titles_fg': '#d699b6',           # purple - Section titles
            'labels_fg': '#dbbc7f',           # yellow - Form labels
            'hover_bg': '#689974',            # aqua - Hover state
            'not_loaded_excel_fg': '#859289', # grey1 - Muted text
            'loaded_excel_fg': '#859966',     # green - Success state
            'scrollbar_bg': '#475258',        # bg3 - Scrollbar
            'field_border': '#425047',         # bg_green - Borders
            'hover_bg_2':   '#9da9d0',           # light-grey - Hover state
        }

        self.configure(bg_color=self.colors['window_bg'])
        self.configure(bg=self.colors['window_bg'])
        # Use a full-window CTkFrame for background
        self.bg_frame = ctk.CTkFrame(self, fg_color=self.colors['window_bg'])
        self.bg_frame.pack(fill="both", expand=True)

        self.excel_file = None
        self.recipients = []
        self.workbook = None
        self.sheet_names = []

        self.setup_ui()

    def setup_ui(self):
        # Excel File Selection
        self.file_label = ctk.CTkLabel(self.bg_frame, text="No file selected", text_color=self.colors['not_loaded_excel_fg'], bg_color=self.colors['window_bg'])
        self.file_label.pack(pady=(10, 5))
        self.browse_btn = ctk.CTkButton(self.bg_frame, text="Browse Excel", command=self.load_excel,
                                        fg_color=self.colors['browse_excel_btn_bg'],
                                        text_color=self.colors['browse_excel_btn_fg'],
                                        hover_color=self.colors['hover_bg'])
        self.browse_btn.pack()

        # Sheet Selection
        self.sheet_var = ctk.StringVar()
        self.sheet_dropdown = ctk.CTkOptionMenu(self.bg_frame, variable=self.sheet_var, values=[],
                                                fg_color=self.colors['dropdown_bg'],
                                                text_color=self.colors['dropdown_fg'],
                                                button_color=self.colors['dropdown_bg'],
                                                button_hover_color=self.colors['hover_bg'])
        self.sheet_dropdown.pack(pady=5)
        self.load_sheet_button = ctk.CTkButton(self.bg_frame, text="Load Sheet", command=self.load_sheet, state="disabled",
                               fg_color=self.colors['send_btn_bg'],
                               text_color=self.colors['send_btn_fg'],
                               hover_color=self.colors['hover_bg'])
        self.load_sheet_button.pack()

        # Gmail Credentials
        self.email_entry = ctk.CTkEntry(self.bg_frame, placeholder_text="Gmail Email",
                        fg_color=self.colors['field_bg'],
                        text_color=self.colors['field_fg'],
                        width=350)
        self.email_entry.pack(pady=5)
        self.password_entry = ctk.CTkEntry(self.bg_frame, placeholder_text="App Password", show="*",
                           fg_color=self.colors['field_bg'],
                           text_color=self.colors['field_fg'],
                           width=350)
        self.password_entry.pack(pady=5)

        # Subject
        today = datetime.today()
        first_day_this_month = today.replace(day=1)
        last_month = first_day_this_month - timedelta(days=1)
        month_year = last_month.strftime("%m.%Y")
        self.subject_entry = ctk.CTkEntry(self.bg_frame, placeholder_text="Email Subject",
                          fg_color=self.colors['field_bg'],
                          text_color=self.colors['field_fg'],
                          width=500)
        self.subject_entry.insert(0, f"Ливче од пресметка на плата за {month_year}")
        self.subject_entry.pack(pady=5)

        # Message Body
        self.message_text = ctk.CTkTextbox(self.bg_frame, height=120,
                           fg_color=self.colors['field_bg'],
                           text_color=self.colors['field_fg'],
                           width=500)
        self.message_text.insert("1.0", f"Почитуван {{name}},\n\nВо прилог ти праќам ливче од пресметка на плата за {month_year}\n\nР.Ѕ. Доколку воочите дека има нејаснотии во ливчето задолжително јавете се во Благајна или на емаил takidaskalo.blagajnik@gmail.com")
        self.message_text.pack(pady=5)

        # Send Button
        self.send_button = ctk.CTkButton(self.bg_frame, text="Send Emails", command=self.send_emails,
                                        fg_color=self.colors['send_btn_bg'],
                                        text_color=self.colors['send_btn_fg'],
                                        hover_color=self.colors['hover_bg'])
        self.send_button.pack(pady=10)

        # Status Log
        self.log_text = ctk.CTkTextbox(self.bg_frame, height=180, state="normal",
                           fg_color=self.colors['status_log_bg'],
                           text_color=self.colors['status_log_fg'],
                           width=500)
        self.log_text.pack(pady=5)

    def log(self, message):
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.update()

    def load_excel(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not file_path:
            return
        try:
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            self.sheet_names = self.workbook.sheetnames
            self.excel_file = file_path
            self.sheet_dropdown.configure(values=self.sheet_names)
            self.sheet_var.set(self.sheet_names[0])
            self.load_sheet_button.configure(state="normal")
            self.file_label.configure(text=f"Loaded: {os.path.basename(file_path)} ({len(self.sheet_names)} sheets)")
            self.log(f"Excel file loaded: {os.path.basename(file_path)}")
            self.log(f"Available sheets: {', '.join(self.sheet_names)}")
            self.log("Please select a sheet and click 'Load Sheet'")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel file:\n{str(e)}")

    def load_sheet(self):
        selected_sheet = self.sheet_var.get()
        if not selected_sheet:
            messagebox.showwarning("Warning", "Please select a sheet!")
            return
        try:
            sheet = self.workbook[selected_sheet]
            self.recipients = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] and row[2]:
                    self.recipients.append({
                        'name': row[0],
                        'email': row[1],
                        'attachment': row[2]
                    })
            self.file_label.configure(text=f"Sheet '{selected_sheet}': {len(self.recipients)} recipients loaded")
            self.log(f"Loaded {len(self.recipients)} recipients from sheet '{selected_sheet}'")
            if len(self.recipients) == 0:
                messagebox.showwarning("Warning", "No recipients found in this sheet!\n\nMake sure the sheet has:\n- Row 1: Headers (Name, Email, AttachmentPath)\n- Row 2+: Data")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sheet:\n{str(e)}")

    def send_emails(self):
        if not self.recipients:
            messagebox.showwarning("Warning", "Please load an Excel file first!")
            return
        sender_email = self.email_entry.get().strip()
        sender_password = self.password_entry.get().strip()
        subject = self.subject_entry.get().strip()
        message_body = self.message_text.get("1.0", "end").strip()
        if not sender_email or not sender_password:
            messagebox.showwarning("Warning", "Please enter Gmail credentials!")
            return
        if not subject or not message_body:
            messagebox.showwarning("Warning", "Please enter subject and message!")
            return
        if not messagebox.askyesno("Confirm", f"Send {len(self.recipients)} emails?"):
            return
        self.send_button.configure(state="disabled")
        self.log("Starting email sending process...")
        success_count = 0
        fail_count = 0
        try:
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(sender_email, sender_password)
            self.log("Connected to Gmail successfully")
            for recipient in self.recipients:
                try:
                    msg = MIMEMultipart()
                    msg['From'] = sender_email
                    msg['To'] = recipient['email']
                    msg['Subject'] = subject
                    personalized_message = message_body.replace("{name}", recipient['name'])
                    msg.attach(MIMEText(personalized_message, 'plain'))
                    if os.path.exists(recipient['attachment']):
                        attachment_path = recipient['attachment']
                        filename = os.path.basename(attachment_path)
                        with open(attachment_path, 'rb') as file:
                            part = MIMEBase('application', 'octet-stream')
                            part.set_payload(file.read())
                            encoders.encode_base64(part)
                            part.add_header('Content-Disposition', 'attachment', filename=filename)
                            msg.attach(part)
                    else:
                        self.log(f"⚠ Warning: Attachment not found for {recipient['name']}: {recipient['attachment']}")
                    server.send_message(msg)
                    success_count += 1
                    self.log(f"✓ Sent to {recipient['name']} ({recipient['email']})")
                except Exception as e:
                    fail_count += 1
                    self.log(f"✗ Failed to send to {recipient['name']}: {str(e)}")
            server.quit()
            self.log(f"\nCompleted! Success: {success_count}, Failed: {fail_count}")
            messagebox.showinfo("Complete", f"Emails sent!\nSuccess: {success_count}\nFailed: {fail_count}")
        except Exception as e:
            self.log(f"Error: {str(e)}")
            messagebox.showerror("Error", f"Failed to connect to Gmail:\n{str(e)}\n\nMake sure you're using an App Password, not your regular password.")
        finally:
            self.send_button.configure(state="normal")

if __name__ == "__main__":
    app = EmailSenderApp()
    app.mainloop()
